from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from clientradar.models.lead import Lead


class ExcelExporter:
    _HEADERS = [
        "#", "Název", "Kategorie", "Město", "Telefon", "Web",
        "Email", "Hodnocení", "Recenze", "Status", "Poznámky", "Datum scrapování",
    ]

    _HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    _HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    _EVEN_FILL = PatternFill(start_color="f0f0f5", end_color="f0f0f5", fill_type="solid")
    _ODD_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    _THIN_BORDER = Border(
        left=Side(style="thin", color="cccccc"),
        right=Side(style="thin", color="cccccc"),
        top=Side(style="thin", color="cccccc"),
        bottom=Side(style="thin", color="cccccc"),
    )

    def export(self, leads: list[Lead], filepath: str) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "ClientRadar Leads"

        for col_idx, header in enumerate(self._HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self._HEADER_FILL
            cell.font = self._HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self._THIN_BORDER

        ws.row_dimensions[1].height = 28

        for row_idx, lead in enumerate(leads, start=2):
            values = [
                row_idx - 1,
                lead.name,
                lead.category,
                lead.city,
                lead.phone,
                lead.website,
                lead.email,
                lead.rating,
                lead.review_count,
                lead.status.value,
                lead.notes,
                lead.scraped_at.strftime("%d.%m.%Y %H:%M") if lead.scraped_at else "",
            ]
            fill = self._EVEN_FILL if row_idx % 2 == 0 else self._ODD_FILL
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = fill
                cell.border = self._THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in (11,)))

        self._auto_column_widths(ws)
        ws.freeze_panes = "A2"
        wb.save(filepath)

    def _auto_column_widths(self, ws) -> None:
        for col_cells in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                val = str(cell.value) if cell.value else ""
                max_length = max(max_length, len(val))
            adjusted = min(max_length + 4, 50)
            ws.column_dimensions[col_letter].width = adjusted
